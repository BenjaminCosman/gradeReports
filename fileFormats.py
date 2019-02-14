import pyexcel as pe
import openpyxl

__all__ = ['getRows']

# Dispatches to one of the functions below
def getRows(sourcePath, isRoster, sheetName):
    ext = sourcePath.suffix
    if ext == '.csv':
        if isRoster:
            return getRowsRosterCSV(sourcePath)
        else:
            return getRowsNormalCSV(sourcePath)
    elif ext == '.xlsx':
        if isRoster:
            if sheetName == None:
                raise Exception("must specify sheetName for roster")
            else:
                return getRowsRosterSingleSheetXLSX(sourcePath, sheetName)
        else:
            if sheetName == None:
                return getRowsNormalMultiSheetXLSX(sourcePath)
            else:
                return getRowsNormalSingleSheetXLSX(sourcePath, sheetName)
    else:
        raise Exception(f"unknown filetype: {str(sourcePath)}")

def getRowsNormalCSV(sourcePath):
    return pe.get_records(file_name=str(sourcePath), auto_detect_float=False, auto_detect_int=False, auto_detect_datetime=False)

def getRowsRosterCSV(sourcePath):
    contents = sourcePath.read_text()
    idx = contents.index("Sec ID,PID,Student,Credits,College,Major,Level,Email\n")
    contents = contents[idx:]
    rows = pe.get_records(file_content=contents, file_type='csv', auto_detect_float=False, auto_detect_int=False, auto_detect_datetime=False)
    return rows

def getRowsNormalSingleSheetXLSX(sourcePath, sheetname):
    wb = openpyxl.load_workbook(filename=str(sourcePath), data_only=True)
    ws = wb[sheetname]
    return pe.get_records(array=[[str(x) for x in row] for row in ws.values])
    # rows = pe.get_records(file_name=str(sourcePath), sheet_name=sheetname, auto_detect_float=False, auto_detect_int=False, auto_detect_datetime=False)
    # if '' in rows[0].keys():
    #     #TODO: remove these once issue is resolved (https://github.com/pyexcel/pyexcel/issues/170)
    #     raise Exception(f"cannot have blank column headers in xlsx file (sheet {sheetname} of {sourcePath})")
    # return rows

def getRowsNormalMultiSheetXLSX(sourcePath):
    wb = openpyxl.load_workbook(filename=str(sourcePath), data_only=True)
    output = []
    for sheetname in wb.sheetnames:
        rows = pe.get_records(array=[[str(x) for x in row] for row in wb[sheetname].values])
        for row in rows:
            row.update({"_sheetName": sheetname})
        output += rows
    return output
    # #TODO: probably terrible performance - loading whole book just to read sheet names
    # sheets = pe.get_book_dict(file_name=str(sourcePath)).keys()
    # output = []
    # for sheetname in sheets:
    #     rows = pe.get_records(file_name=str(sourcePath), sheet_name=sheetname, auto_detect_float=False, auto_detect_int=False, auto_detect_datetime=False)
    #     if '' in rows[0].keys():
    #         raise Exception(f"cannot have blank column headers in xlsx file (sheet {sheetname} of {sourcePath})")
    #     for row in rows:
    #         row.update({"_sheetName": sheetname})
    #     output += rows
    # return output

def getRowsRosterSingleSheetXLSX(sourcePath, sheetname):
    allRows = pe.get_array(file_name=str(sourcePath), sheet_name=sheetname, auto_detect_float=False, auto_detect_int=False, auto_detect_datetime=False)
    idx = allRows.index(['Sec ID','PID','Student','Credits','College','Major','Level','Email'])
    if idx == -1:
        raise Exception("not a roster")
    return pe.get_records(array=allRows[idx:], auto_detect_float=False, auto_detect_int=False, auto_detect_datetime=False)
